# -*- coding: utf-8 -*-
"""
Created on Sat Apr 28 22:23:58 2018

@author: Elias
"""

from openpyxl import Workbook
from openpyxl import load_workbook
#from pandas import Series
#from pandas import Index
#from pandas import DataFrame

#import numpy as np
#import os
#import pandas as pd

#把旧BOM里面的BYD料号读取出来
#global bom_old,bom_old_sheet,bom_old_sheet_len
bom_old = load_workbook("DANA_MB_BOM_EVT_SKU3_i7_SMT_v1.8_Final_20180502.xlsx")
#bom_old_sheet = bom_old.get_sheet_by_name('Sheet1') 系统不建议的语法
bom_old_sheet = bom_old['Sheet1'] #系统建议语法
bom_old_sheet_len = len(bom_old_sheet['A'])
old_content = []
for i in range(bom_old_sheet_len):
    old_content.append(bom_old_sheet.cell(row=i+2,column=2).value)

#把新BOM里面的BYD料号读取出来
bom_new = load_workbook("DANA_MB_BOM_EVT_SKU1_SCH导出0422.xlsx")
#bom_new_sheet = bom_new.get_sheet_by_name('Sheet1')
bom_new_sheet = bom_new['Sheet1']
bom_new_sheet_len = len(bom_old_sheet['A'])
new_content = []
for j in range(bom_new_sheet_len):
    new_content.append(bom_new_sheet.cell(row=j+2,column=2).value)

#新建一个表格，把旧新BOM里的BYD料号分别放到sheet1的A列和B列
bom_byd = Workbook()
#bom_byd_sheet1 = bom_byd.get_sheet_by_name('Sheet')
bom_byd_sheet1 = bom_byd['Sheet']
for n in range(bom_old_sheet_len):
    bom_byd_sheet1.cell(row=n+1,column=1).value = old_content[n]
    bom_byd_sheet1.cell(row=n+1,column=2).value = new_content[n]

#求旧BOM和新BOM都有的BYD料号
bom_common = []
bom_common = list(set(old_content) & set(new_content))
m = len(bom_common)
#将都有的BYD料号放到sheet1的C列
for p in range(1,m):
    bom_byd_sheet1.cell(row=p,column=3).value = bom_common[p-1]
    
#保存表格
bom_byd.save(filename="BOM_BYD.xlsx")

bom_old = load_workbook("DANA_MB_BOM_EVT_SKU3_i7_SMT_v1.8_Final_20180502.xlsx")
bom_old_sheet = bom_old['Sheet1']
bom_old_len = len(bom_old_sheet['A'])
#把共同BYD PN拿出来
for t in range(m):    
    bydpn = bom_common[t]
    #遍历旧BOM里的BYD PN，找到有共同BYD PN的位置，得到该位置前一列的序号
    for q in range(bom_old_sheet_len):     
        if(bom_old_sheet.cell(row = q+2, column = 2).value == bydpn):
            qwe = bom_old_sheet.cell(row = q+2, column = 2).value
            #bom_old_sheet.cell(row = q+2, column = 13).value = "true"
            
            item = bom_old_sheet.cell(row = q+2, column = 1).value         
            #如果相同序号有2个以上，判定为这颗料有替代料
            for r in range(bom_old_len):
                

                if(bom_old_sheet.cell(row = r+2,column = 1 ).value == item):
                    qwer = bom_old_sheet.cell(row = r+2,column = 1 ).value
                    bom_old_sheet.cell(row = r+2, column = 13).value = q
                    #遍历BOM里有相同序号的行，将后面的第10列标为true
                    r = 0
      
print("Done, 筛选第13列的true即可提取出新BOM与旧BOM有相同部分的主料和替代料信息")

bom_old.save(filename = "DANA_MB_BOM_EVT_SKU3_i7_SMT_v1.8_Final_20180502.xlsx")


'''
#将旧BOM中有替代料的序号列出来
data = pd.read_excel('DANA1_OLD.xlsx','Sheet1')

item_common = data.loc[data['BYDPN'] == '100000-00', 'item']
item_common_convert = item_common[0]
item_common_1 = data.loc[data['item'] == '10','Description']
#replace_info = data.loc[data['item'] == 'item_common_convert[0]']
    #data[data['item']==df_item[p]].head()
 
    #bom_byd_sheet1.cell(row=q+1,column=4).value = df_item[q]


print(item_common)
print(item_common_convert)
#print(item_common_1)

#print(data)

#x = Series()
#for q in range(0,m-1):
#print(df)

 '''   





